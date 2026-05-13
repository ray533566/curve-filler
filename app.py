import io
import os
import re
import zipfile
import urllib.request

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

# ─── Page config ────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="2xFR4 Curve Filler",
    page_icon="📡",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─── GitHub template URL ─────────────────────────────────────────────────────
# Replace with the raw URL from your GitHub repository, e.g.:
# https://raw.githubusercontent.com/johndoe/curve-filler/main/TEC_2xFR4_Curve_PY.xlsx
TEMPLATE_URL = (
    TEMPLATE_URL = "https://raw.githubusercontent.com/ray533566/curve-filler/main/TEC_2xFR4_Curve_PY.xlsx"
)

# ─── Custom CSS ─────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Space+Mono:wght@400;700&family=DM+Sans:wght@300;400;500;600&display=swap');
html, body, [class*="css"] { font-family: 'DM Sans', sans-serif; }
.stApp { background: #0e1117; color: #e2e8f0; }
section[data-testid="stSidebar"] { background: #161b27 !important; border-right: 1px solid #2d3748; }
h1 { font-family: 'Space Mono', monospace !important; color: #63b3ed !important; letter-spacing: -0.02em; }
h2, h3 { font-family: 'Space Mono', monospace !important; color: #90cdf4 !important; }
.card { background: #161b27; border: 1px solid #2d3748; border-radius: 10px; padding: 1.2rem 1.5rem; margin-bottom: 1rem; }
.card-title { font-family: 'Space Mono', monospace; font-size: 0.78rem; letter-spacing: 0.12em; text-transform: uppercase; color: #63b3ed; margin-bottom: 0.5rem; }
[data-testid="stFileUploader"] { background: #1a2035 !important; border: 1.5px dashed #4a5568 !important; border-radius: 8px !important; }
.stButton > button { background: linear-gradient(135deg, #2b6cb0, #3182ce) !important; color: white !important; border: none !important; border-radius: 8px !important; font-family: 'Space Mono', monospace !important; font-size: 0.85rem !important; padding: 0.6rem 1.4rem !important; transition: all 0.2s ease !important; }
.stButton > button:hover { background: linear-gradient(135deg, #3182ce, #4299e1) !important; transform: translateY(-1px); box-shadow: 0 4px 15px rgba(66, 153, 225, 0.35) !important; }
[data-testid="stDownloadButton"] > button { background: #1a3a2a !important; border: 1px solid #276749 !important; color: #68d391 !important; border-radius: 6px !important; font-size: 0.78rem !important; padding: 0.35rem 0.9rem !important; width: 100%; font-family: 'Space Mono', monospace !important; }
[data-testid="stDownloadButton"] > button:hover { background: #22543d !important; border-color: #48bb78 !important; }
.pill { display: inline-block; padding: 2px 10px; border-radius: 20px; font-size: 0.72rem; font-family: 'Space Mono', monospace; font-weight: 700; letter-spacing: 0.06em; }
.pill-pc   { background: #1a365d; color: #63b3ed; border: 1px solid #2b6cb0; }
.pill-wrp  { background: #322659; color: #b794f4; border: 1px solid #553c9a; }
.info-box { background: #1a2744; border-left: 3px solid #63b3ed; border-radius: 0 6px 6px 0; padding: 0.6rem 1rem; margin: 0.4rem 0; font-size: 0.85rem; color: #bee3f8; }
.warn-box { background: #2d1b00; border-left: 3px solid #f6ad55; border-radius: 0 6px 6px 0; padding: 0.6rem 1rem; margin: 0.4rem 0; font-size: 0.85rem; color: #fbd38d; }
.success-box { background: #1c3a2a; border-left: 3px solid #68d391; border-radius: 0 6px 6px 0; padding: 0.6rem 1rem; margin: 0.4rem 0; font-size: 0.85rem; color: #9ae6b4; }
hr { border-color: #2d3748 !important; }
</style>
""", unsafe_allow_html=True)

# ─── Constants ───────────────────────────────────────────────────────────────
CH_SHEETS = [
    "1_Operational", "2_Operational", "3_Operational", "4_Operational",
    "1_Maximum",     "2_Maximum",     "3_Maximum",     "4_Maximum",
]

PC_COL_MAP = {
    "[CHNumber]": 13, "[CH_Pass_Fail]": 14, "[FailureCodeID]": 15,
    "[Times]": 16, "[Case_Temp]": 17, "[Wavelength]": 18, "[dW/dT]": 19,
    "[TXP]": 20, "[Peak_WL_Jump_Max]": 21, "[Peak_WL_Jump_Avg]": 22,
    "[Peak_WL_Jump_Min]": 23,
    "[CH1_DDMI_Bias]": 24, "[CH2_DDMI_Bias]": 25,
    "[CH3_DDMI_Bias]": 26, "[CH4_DDMI_Bias]": 27,
    "[CH5_DDMI_Bias]": 28, "[CH6_DDMI_Bias]": 29,
    "[CH7_DDMI_Bias]": 30, "[CH8_DDMI_Bias]": 31,
}

WRP_COL_MAP = {
    "TESTNUMBER": 1, "WO": 2, "TESTSN": 3, "OPERATION": 4, "PN": 5,
    "EQUPMENT": 6, "FIXTUREID": 7, "TESTRESULT": 8, "TESTOP": 9,
    "TESTDATE": 10, "TESTREASON": 11, "TEST_STATUS": 12,
    "CHNumber": 13, "CH_Pass_Fail": 14, "FailureCodeID": 15,
    "Times": 16, "Case_Temp": 17, "Wavelength": 18, "dW/dT": 19,
    "TXP": 20, "Peak_WL_Jump_Max": 21, "Peak_WL_Jump_Avg": 22,
    "Peak_WL_Jump_Min": 23,
    "CH1_DDMI_Bias": 24, "CH2_DDMI_Bias": 25,
    "CH3_DDMI_Bias": 26, "CH4_DDMI_Bias": 27,
    "CH5_DDMI_Bias": 28, "CH6_DDMI_Bias": 29,
    "CH7_DDMI_Bias": 30, "CH8_DDMI_Bias": 31,
}

# ─── Template loader (cached) ─────────────────────────────────────────────────

@st.cache_data(show_spinner=False)
def load_template_from_github(url: str) -> bytes:
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=30) as resp:
        return resp.read()


# ─── Core processing functions ───────────────────────────────────────────────

def increment_formula(formula: str, delta: int) -> str:
    def bump(m):
        return f"{m.group(1)}{int(m.group(2)) + delta}"
    return re.sub(r"([A-Z]+)(\d+)", bump, formula)


def extend_formula_sheet(ws, target_rows: int):
    current_last = ws.max_row
    if target_rows + 1 <= current_last:
        return
    ref_row = 2
    for new_row in range(current_last + 1, target_rows + 2):
        delta = new_row - ref_row
        for col in range(1, ws.max_column + 1):
            ref_val = ws.cell(ref_row, col).value
            if ref_val is None:
                continue
            if isinstance(ref_val, str) and ref_val.startswith("="):
                ws.cell(new_row, col).value = increment_formula(ref_val, delta)
            elif col in (6, 7):
                ws.cell(new_row, col).value = ref_val


def clear_ch_data(wb):
    for sh in CH_SHEETS:
        if sh not in wb.sheetnames:
            continue
        ws = wb[sh]
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.value = None


def write_df_to_sheet(ws, df: pd.DataFrame, col_map: dict):
    active_map = {c: col_map[c] for c in col_map if c in df.columns}
    for ri, (_, row) in enumerate(df.iterrows(), start=2):
        for col, idx in active_map.items():
            val = row[col]
            if val is not None and not (isinstance(val, float) and pd.isna(val)):
                ws.cell(ri, idx).value = val


def build_workbook(template_bytes: bytes, ch_data: dict, col_map: dict):
    wb = load_workbook(io.BytesIO(template_bytes))
    clear_ch_data(wb)

    for sheet_name, df in ch_data.items():
        if sheet_name not in wb.sheetnames or df.empty:
            continue
        write_df_to_sheet(wb[sheet_name], df, col_map)

    op_sheets  = ["1_Operational", "2_Operational", "3_Operational", "4_Operational"]
    max_sheets = ["1_Maximum",     "2_Maximum",     "3_Maximum",     "4_Maximum"]
    op_rows    = max(len(ch_data.get(s, pd.DataFrame())) for s in op_sheets)
    max_rows   = max(len(ch_data.get(s, pd.DataFrame())) for s in max_sheets)
    curve_rows = max(op_rows, max_rows)

    extend_formula_sheet(wb["Operational"], op_rows)
    extend_formula_sheet(wb["Maximum"],     max_rows)
    extend_formula_sheet(wb["Curve"],       curve_rows)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), op_rows, max_rows


def read_csv_safe(file_obj) -> pd.DataFrame:
    raw = file_obj.read()
    for enc in ("utf-8-sig", "utf-8", "cp950", "latin-1"):
        try:
            return pd.read_csv(io.BytesIO(raw), encoding=enc)
        except Exception:
            continue
    return pd.DataFrame()


def extract_sn_from_filename(fname: str) -> str:
    m = re.search(r"(P\d{12})", fname)
    return m.group(1) if m else os.path.splitext(os.path.basename(fname))[0]


def process_pc_csv_pair(op_df: pd.DataFrame, max_df: pd.DataFrame) -> dict:
    ch_data = {}
    for s in ["1_Operational", "2_Operational", "3_Operational", "4_Operational"]:
        ch_data[s] = op_df[op_df["[CHNumber]"] == s].reset_index(drop=True)
    for s in ["1_Maximum", "2_Maximum", "3_Maximum", "4_Maximum"]:
        ch_data[s] = max_df[max_df["[CHNumber]"] == s].reset_index(drop=True)
    return ch_data


def process_wrp_csv(df: pd.DataFrame) -> list:
    groups = df["TESTNUMBER"].unique().tolist()
    return [(i, df[df["TESTNUMBER"] == g].reset_index(drop=True)) for i, g in enumerate(groups)]


def wrp_group_to_ch_data(group_df: pd.DataFrame) -> dict:
    return {s: group_df[group_df["CHNumber"] == s].reset_index(drop=True) for s in CH_SHEETS}


# ─── UI ──────────────────────────────────────────────────────────────────────

def sidebar():
    with st.sidebar:
        st.markdown("## 📡 2xFR4 Curve Filler")
        st.markdown("---")
        st.markdown("""
<div class='card'>
<div class='card-title'>How to use</div>

**Step 1** — Upload PC Raw CSVs (zip or multiple files)<br>
**Step 2** — Upload WRP Raw CSVs (zip or multiple files)<br>
**Step 3** — Click <b>Process</b> and download results<br><br>
<span style='color:#63b3ed;font-size:0.8rem'>📄 Template is auto-loaded from GitHub.</span>
</div>
""", unsafe_allow_html=True)

        st.markdown("""
<div class='card'>
<div class='card-title'>WRP Grouping Rule</div>
Each unique <code>TESTNUMBER</code> → separate Excel file<br><br>
<span class='pill pill-wrp'>Group 1</span> → <code>SN_WRP.xlsx</code><br><br>
<span class='pill pill-wrp'>Group 2</span> → <code>SN_WRP_1.xlsx</code><br><br>
<span class='pill pill-wrp'>Group 3</span> → <code>SN_WRP_2.xlsx</code>
</div>
""", unsafe_allow_html=True)

        st.markdown("""
<div class='card'>
<div class='card-title'>Supported Formats</div>
Raw data: <code>.csv</code> or <code>.zip</code>
</div>
""", unsafe_allow_html=True)


def main():
    sidebar()

    st.markdown("# 2xFR4 Curve Filler")
    st.markdown("Automatically fill raw test data into Excel curve templates — PC & WRP, multi-group aware.")
    st.markdown("---")

    # ── Template status ────────────────────────────────────────────────────
    template_bytes = None
    try:
        template_bytes = load_template_from_github(TEMPLATE_URL)
        st.markdown(
            "<div class='success-box'>✓ Template <code>TEC_2xFR4_Curve_PY.xlsx</code> loaded from GitHub</div>",
            unsafe_allow_html=True,
        )
    except Exception as e:
        st.markdown(
            f"<div class='warn-box'>⚠️ Could not load template from GitHub: {e}<br>"
            "Please update <code>TEMPLATE_URL</code> in <code>app.py</code> with your raw GitHub URL.</div>",
            unsafe_allow_html=True,
        )

    st.markdown("")

    col1, col2 = st.columns(2)

    with col1:
        st.markdown("### 🖥️ PC Raw Data")
        pc_files = st.file_uploader(
            "Upload PC CSVs or ZIP",
            type=["csv", "zip"],
            accept_multiple_files=True,
            key="pc",
            label_visibility="collapsed",
        )
        if pc_files:
            st.markdown(
                f"<div class='info-box'><span class='pill pill-pc'>PC</span>&nbsp; {len(pc_files)} file(s) uploaded</div>",
                unsafe_allow_html=True,
            )

    with col2:
        st.markdown("### 📡 WRP Raw Data")
        wrp_files = st.file_uploader(
            "Upload WRP CSVs or ZIP",
            type=["csv", "zip"],
            accept_multiple_files=True,
            key="wrp",
            label_visibility="collapsed",
        )
        if wrp_files:
            st.markdown(
                f"<div class='info-box'><span class='pill pill-wrp'>WRP</span>&nbsp; {len(wrp_files)} file(s) uploaded</div>",
                unsafe_allow_html=True,
            )

    st.markdown("---")
    process_btn = st.button("⚡  Process & Generate Excel Files", use_container_width=True)

    if process_btn:
        if template_bytes is None:
            try:
                with st.spinner("Downloading template from GitHub…"):
                    template_bytes = load_template_from_github(TEMPLATE_URL)
            except Exception as e:
                st.error(
                    f"❌ Failed to download template from GitHub: {e}\n\n"
                    "Please update `TEMPLATE_URL` in `app.py` to point to your raw GitHub file."
                )
                return

        if not pc_files and not wrp_files:
            st.warning("Please upload at least one PC or WRP raw data file.")
            return

        results = []
        log_lines = []
        progress = st.progress(0, text="Starting…")

        def load_csvs_from_uploads(uploads):
            out = []
            for uf in uploads:
                if uf.name.lower().endswith(".zip"):
                    with zipfile.ZipFile(io.BytesIO(uf.read())) as z:
                        for name in z.namelist():
                            if name.lower().endswith(".csv") and not name.startswith("__"):
                                with z.open(name) as f:
                                    raw = f.read()
                                    for enc in ("utf-8-sig", "utf-8", "cp950", "latin-1"):
                                        try:
                                            df = pd.read_csv(io.BytesIO(raw), encoding=enc)
                                            out.append((os.path.basename(name), df))
                                            break
                                        except Exception:
                                            continue
                else:
                    uf.seek(0)
                    df = read_csv_safe(uf)
                    out.append((uf.name, df))
            return out

        # ── PC processing ─────────────────────────────────────────────────
        if pc_files:
            progress.progress(10, text="Loading PC raw data…")
            pc_csvs = load_csvs_from_uploads(pc_files)

            pc_by_sn: dict = {}
            for fname, df in pc_csvs:
                if df.empty:
                    continue
                sn = extract_sn_from_filename(fname)
                pc_by_sn.setdefault(sn, {})
                if "Operational" in fname:
                    pc_by_sn[sn]["op"] = df
                elif "Maximum" in fname:
                    pc_by_sn[sn]["max"] = df

            total_pc = len(pc_by_sn)
            for idx, (sn, pair) in enumerate(pc_by_sn.items()):
                progress.progress(10 + int(30 * idx / max(total_pc, 1)), text=f"PC: {sn}…")
                if "op" not in pair or "max" not in pair:
                    log_lines.append(f"⚠️ PC {sn}: missing Operational or Maximum CSV — skipped")
                    continue
                ch_data = process_pc_csv_pair(pair["op"], pair["max"])
                wb_bytes, op_r, max_r = build_workbook(template_bytes, ch_data, PC_COL_MAP)
                fname_out = f"TEC_2xFR4_Curve_{sn}_PC.xlsx"
                results.append((fname_out, wb_bytes, "PC"))
                log_lines.append(f"✓ {fname_out}  (Op={op_r} rows, Max={max_r} rows)")

        # ── WRP processing ────────────────────────────────────────────────
        if wrp_files:
            progress.progress(45, text="Loading WRP raw data…")
            wrp_csvs = load_csvs_from_uploads(wrp_files)

            wrp_by_sn: dict = {}
            for fname, df in wrp_csvs:
                if df.empty:
                    continue
                sn = extract_sn_from_filename(fname)
                if sn in wrp_by_sn:
                    wrp_by_sn[sn] = pd.concat([wrp_by_sn[sn], df], ignore_index=True)
                else:
                    wrp_by_sn[sn] = df

            total_wrp = sum(
                len(df["TESTNUMBER"].unique())
                for df in wrp_by_sn.values()
                if "TESTNUMBER" in df.columns
            )
            done = 0
            for sn, df in wrp_by_sn.items():
                if "TESTNUMBER" not in df.columns:
                    log_lines.append(f"⚠️ WRP {sn}: no TESTNUMBER column — skipped")
                    continue
                groups = process_wrp_csv(df)
                for group_idx, group_df in groups:
                    progress.progress(
                        45 + int(50 * done / max(total_wrp, 1)),
                        text=f"WRP: {sn} group {group_idx + 1}/{len(groups)}…",
                    )
                    ch_data = wrp_group_to_ch_data(group_df)
                    wb_bytes, op_r, max_r = build_workbook(template_bytes, ch_data, WRP_COL_MAP)
                    suffix = "WRP" if group_idx == 0 else f"WRP_{group_idx}"
                    fname_out = f"TEC_2xFR4_Curve_{sn}_{suffix}.xlsx"
                    results.append((fname_out, wb_bytes, "WRP"))
                    tn = group_df["TESTNUMBER"].iloc[0]
                    log_lines.append(f"✓ {fname_out}  (Op={op_r}, Max={max_r}) TESTNUMBER={tn}")
                    done += 1

        progress.progress(100, text="Done!")

        # ── Results display ───────────────────────────────────────────────
        st.markdown("---")
        st.markdown(f"### ✅ Generated {len(results)} file(s)")

        with st.expander("📋 Processing log", expanded=False):
            for line in log_lines:
                color = "success-box" if line.startswith("✓") else "warn-box"
                st.markdown(f"<div class='{color}'>{line}</div>", unsafe_allow_html=True)

        if not results:
            st.warning("No output files were generated. Check your uploads.")
            return

        pc_results  = [(n, b) for n, b, t in results if t == "PC"]
        wrp_results = [(n, b) for n, b, t in results if t == "WRP"]

        tab_pc, tab_wrp, tab_all = st.tabs([
            f"🖥️ PC ({len(pc_results)})",
            f"📡 WRP ({len(wrp_results)})",
            "📦 Download All as ZIP",
        ])

        def download_grid(items, pill_class):
            cols = st.columns(3)
            for i, (fname, data) in enumerate(items):
                with cols[i % 3]:
                    label = fname.replace("TEC_2xFR4_Curve_", "").replace(".xlsx", "")
                    st.markdown(
                        f"<div style='text-align:center;margin-bottom:4px'>"
                        f"<span class='pill {pill_class}'>{label}</span></div>",
                        unsafe_allow_html=True,
                    )
                    st.download_button(
                        label="⬇  Download",
                        data=data,
                        file_name=fname,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"dl_{fname}",
                    )

        with tab_pc:
            if pc_results:
                download_grid(pc_results, "pill-pc")
            else:
                st.info("No PC files generated.")

        with tab_wrp:
            if wrp_results:
                download_grid(wrp_results, "pill-wrp")
            else:
                st.info("No WRP files generated.")

        with tab_all:
            zip_buf = io.BytesIO()
            with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
                for fname, data, _ in results:
                    zf.writestr(fname, data)
            st.download_button(
                label="📦  Download All Files (.zip)",
                data=zip_buf.getvalue(),
                file_name="TEC_2xFR4_Curve_Output.zip",
                mime="application/zip",
                use_container_width=True,
            )
            st.markdown(
                f"<div class='info-box'>Contains {len(results)} Excel files — "
                f"{len(pc_results)} PC + {len(wrp_results)} WRP</div>",
                unsafe_allow_html=True,
            )


if __name__ == "__main__":
    main()
